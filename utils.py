import json
import pandas as pd
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from mysql.connector import Error


def process_excel(file_path, input_sheet_name):
    print(f"📂 Loading data from: {file_path}")
    try:
        df = pd.read_excel(file_path, sheet_name=input_sheet_name)
    except FileNotFoundError:
        print(f"❌ Error: File '{file_path}' not found!")
        return
    except ValueError:
        print(f"❌ Error: Sheet '{input_sheet_name}' not found in '{file_path}'!")
        return
    
    df.columns = df.columns.str.strip()
    print("✅ Data Loaded Successfully!")
    
    if df.empty:
        print("❌ Error: Input sheet is empty!")
        return
    
    book = load_workbook(file_path)
    for sheet in ["Country_Total", "Month_Average"]:
        if sheet in book.sheetnames:
            del book[sheet]
    book.save(file_path)
    
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        country_groups = df.groupby("Country")
        for country, group in country_groups:
            sheet_name = str(country)[:31]
            print(f"📊 Writing data for country: {country}")
            group.to_excel(writer, sheet_name=sheet_name, index=False)
        
        country_totals = country_groups.agg({"Units Sold": 'sum', "Sales": 'sum', "Profit": 'sum'}).reset_index()
        print("📌 Writing 'Country_Total' sheet...")
        country_totals.to_excel(writer, sheet_name="Country_Total", index=False)
    
    print("✅ Data processing completed!")

def apply_alternating_row_colors(file_path, input_sheet_name):
    print(f"🎨 Applying alternating row colors to: {file_path}")
    book = load_workbook(file_path)
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for sheet_name in book.sheetnames:
        if sheet_name != input_sheet_name:
            ws = book[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = blue_fill
    book.save(file_path)
    print("✅ Formatting applied successfully!")


class FoodDataProcessor:
    def __init__(self, json_file, db_config):
        """Initialize with JSON file and database configuration"""
        self.json_file = json_file
        self.db_config = db_config
        self.connection = None

    def connect_to_db(self):
        """Create a connection to the MySQL database"""
        try:
            self.connection = mysql.connector.connect(
                host=self.db_config['host'],
                database=self.db_config['database'],
                user=self.db_config['user']
                #password=self.db_config['password']  # Added password
            )
            if self.connection.is_connected():
                print("Connected to MySQL database")
        except Error as e:
            print(f"Error connecting to MySQL database: {e}")
            self.connection = None  # Ensure connection is set to None if failed
            raise

    def create_table_for_category(self, category, data):
        """Dynamically create a table for each category based on the keys in the data"""
        cursor = None
        try:
            if self.connection is None:
                self.connect_to_db()

            cursor = self.connection.cursor()

            # Create the table with dynamic columns based on the keys of the first item in the data list
            columns = ", ".join([f"{key} VARCHAR(255)" for key in data[0].keys()])

            create_table_query = f'''
                CREATE TABLE IF NOT EXISTS {category} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    {columns}
                )
            '''

            cursor.execute(create_table_query)
            print(f"✅ Table '{category}' created (or exists already)")

        except mysql.connector.Error as e:
            print(f"Error creating table {category}: {e}")
        finally:
            if cursor:
                cursor.close()

    def insert_data_to_table(self, category, data):
        """Insert data into the dynamically created table for each category"""
        cursor = None
        try:
            if self.connection is None:
                self.connect_to_db()

            cursor = self.connection.cursor()

            # Prepare the insert query dynamically based on the keys of the data
            columns = ", ".join(data[0].keys())
            values_placeholder = ", ".join(["%s"] * len(data[0]))

            insert_query = f'''
                INSERT INTO {category} ({columns})
                VALUES ({values_placeholder})
            '''

            for item in data:
                cursor.execute(insert_query, tuple(item.values()))

            self.connection.commit()  # Commit the changes to the database
            print(f"✅ Data successfully inserted into the '{category}' table")

        except mysql.connector.Error as e:
            print(f"Database error while inserting data into {category}: {e}")
        except Exception as e:
            print(f"Error: {e}")
        finally:
            if cursor:
                cursor.close()


    def process_data(self):
        """Process the JSON data, create tables, and insert data into each table"""
        try:
            with open(self.json_file, 'r') as file:
                data = json.load(file)

            # Check if the data is a dictionary (which it should be, with categories as keys)
            if not isinstance(data, dict):
                print("Error: Data in the JSON file is not in the expected format (a dictionary of categories).")
                return

            # Iterate through each category and process it
            for category, items in data.items():
                print(f"Processing category: {category}")
                self.create_table_for_category(category, items)  # Create table for the category
                self.insert_data_to_table(category, items) # Insert data into the category table
                self.write_to_excel()  

            print("✅ All data processed successfully!")

        except Exception as e:
            print(f"Error: {e}")
            
            
    #def process_data(self):
     #   """Process JSON data and write it to the MySQL database and Excel file"""
      #  self.connect_to_db()  # Ensure we are connected to the database
      #  self.create_table_for_category(category, data)  # Create the tables
     #   self.insert_data_to_table(category, data)  # Write data to the database
      #  self.write_to_excel()  # Write data to an Excel file

    def write_to_excel(self):
        """Write JSON data to an Excel file with three sheets (food, topping, batter)"""
        try:
            with open(self.json_file, 'r') as file:
                data = json.load(file)

            # Debugging: Print the loaded data
            print("Loaded data:", data)

            # Ensure data is a dictionary with categories (Food, Topping, Batter)
            if not isinstance(data, dict):
                print("Error: Data is not in the expected format (dictionary of categories).")
                return

            # Separate the data into categories
            food_data = data.get('Food', [])
            topping_data = data.get('Topping', [])
            batter_data = data.get('Batter', [])

            # Convert each category's data to a pandas DataFrame
            food_df = pd.DataFrame(food_data)
            topping_df = pd.DataFrame(topping_data)
            batter_df = pd.DataFrame(batter_data)

            # Write data to an Excel file with three sheets
            with pd.ExcelWriter('food_data.xlsx') as writer:
                food_df.to_excel(writer, sheet_name='food', index=False)
                topping_df.to_excel(writer, sheet_name='topping', index=False)
                batter_df.to_excel(writer, sheet_name='batter', index=False)

            print("✅ Data written to Excel file 'food_data.xlsx'")

        except Exception as e:
            print(f"Error writing to Excel: {e}")
